import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse


def generate_excel_report(data):
    # Создаем новый Excel файл
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Report"

    # Настраиваем ширину колонок
    column_widths = [20, 20, 20, 20]
    for i, width in enumerate(column_widths, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    # Создаем заголовки таблицы
    headers = ["TIME IN", "TIME OUT", "TOTAL TIME", "FRAMES SOLD", "MEDIA USED", "CASH", "CARD", "TOTAL"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Вставляем данные
    for row_num, row_data in enumerate(data, start=2):
        for col_num, cell_value in enumerate(row_data, start=1):
            cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Возвращаем Excel в HTTP ответе
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="report.xlsx"'
    workbook.save(response)
    return response
